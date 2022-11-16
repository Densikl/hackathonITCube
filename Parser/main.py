import openpyxl
import sys
sys.path.insert(1, '/hackathon/Parser/Model/')

def GetNodes(source):
    wb = openpyxl.load_workbook(source)
    sheet = wb.active
    x = []
    y = []
    cities = {}
    count_x = 0
    for row in sheet.iter_rows():
        count_y = 0
        for cell in row:
            if count_x == 0:
                x.append(cell.value)
                if cell.value != None:
                    cities[cell.value] = {}
            if count_y == 0:
                y.append(cell.value)
            count_y+=1
        count_x+=1

    for row in sheet.iter_rows():
        tmp_city = list()
        start_city = str()
        for cell in row:
            if type(cell.value) == type(1):
                start_city = y[cell.row - 1]
                cities[start_city][x[cell.column - 1]] = cell.value
    
    return cities
